import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook
import os


def is_file_extension(name):
    common_file_extensions = ['.txt', '.doc', '.docx', '.xlsx', '.pdf']
    file_extension = os.path.splitext(name)[1].lower()
    return file_extension in common_file_extensions


def select_root_directory():
    root_directory = filedialog.askdirectory()
    root_directory_entry.delete(0, tk.END)
    root_directory_entry.insert(0, root_directory)


def select_excel_file():
    excel_file = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, excel_file)


def create_directories():
    root_directory = root_directory_entry.get()
    excel_file = excel_file_entry.get()

    if root_directory and excel_file:
        try:
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                current_folder = root_directory

                empty_count = 0  # Count of consecutive empty cells

                for cell in row:
                    if empty_count >= 4:
                        break

                    if cell is None:
                        empty_count += 1
                        continue

                    name = str(cell)

                    if is_file_extension(name):
                        file_name, file_extension = os.path.splitext(name)
                        file_path = os.path.join(current_folder, file_name + file_extension)
                        open(file_path, 'w').close()
                        continue

                    folder_path = os.path.join(current_folder, name)
                    os.makedirs(folder_path, exist_ok=True)
                    current_folder = folder_path
                    empty_count = 0

            messagebox.showinfo("成功", "目录结构生成成功！")
        except Exception as e:
            messagebox.showerror("错误", f"发生错误：{str(e)}")
    else:
        messagebox.showwarning("警告", "请先选择根目录和Excel文件！")


# 创建主窗口
root = tk.Tk()
root.title("文件夹批量创建工具")
root.geometry("500x140")

# 设置列权重
root.columnconfigure(1, weight=1)
root.columnconfigure(3, weight=1)

# 第一行
root_directory_label = tk.Label(root, text="根目录：", anchor='w')
root_directory_label.grid(row=0, column=0, sticky='w', padx=10, pady=(10, 5))

root_directory_entry = tk.Entry(root)
root_directory_entry.grid(row=0, column=1, columnspan=3, padx=10, sticky='ew')

select_root_directory_button = tk.Button(root, text="选择根目录", width=15, command=select_root_directory)
select_root_directory_button.grid(row=0, column=4, pady=5)

# 第二行
excel_file_label = tk.Label(root, text="预定义EXCEL：", anchor='w')
excel_file_label.grid(row=1, column=0, sticky='w', padx=10, pady=5)

excel_file_entry = tk.Entry(root)
excel_file_entry.grid(row=1, column=1, columnspan=3, padx=10, sticky='ew')

select_excel_file_button = tk.Button(root, text="选择预定义EXCEL", width=15, command=select_excel_file)
select_excel_file_button.grid(row=1, column=4, pady=5)

# 第三行
create_button = tk.Button(root, text="创建", width=15, command=create_directories)
create_button.grid(row=2, column=0, columnspan=5, pady=(10, 5))

# 运行主循环
root.mainloop()
